#!/usr/bin/env python
# 매물 엑셀(여러 개) → 급매 선별 + 정밀 지오코딩 + Supabase 등록 (자동 연동)
#
# 사용:
#   python scripts/import-listings.py "<엑셀1>" "<엑셀2>" ...          # 미리보기(dry-run)
#   python scripts/import-listings.py "<엑셀들...>" --apply             # Supabase 실제 등록(지역별 교체)
#   (인자 없으면 Downloads 의 강원_*_아파트매매_20260624.xlsx 자동 수집)
#
# --apply 에는 .env.local 의 SUPABASE_SERVICE_ROLE_KEY 필요.
import json, re, os, sys, time, io, csv, hashlib, collections, glob
import urllib.request, urllib.parse, urllib.error
try: sys.stdout.reconfigure(encoding="utf-8")  # 윈도우 cp949 콘솔에서 이모지 출력 크래시 방지
except Exception: pass

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CP = os.path.join(ROOT, "scripts", "output", "complex_prices.csv")
OUT_SQL = os.path.join(ROOT, "scripts", "output", "properties_insert.sql")
OUT_REP = os.path.join(ROOT, "scripts", "output", "_import_report.txt")
GCACHE = os.path.join(ROOT, "scripts", "output", "_geocache.json")
MIN_SAMPLE, MIN_DISC, MAX_DISC, TODAY = 3, 5.0, 40.0, "2026-06-24"

APPLY = "--apply" in sys.argv
FILES = [a for a in sys.argv[1:] if not a.startswith("--")]
if not FILES:
    FILES = sorted(glob.glob(os.path.join(os.path.expanduser("~"), "Downloads", "*_아파트매매_*.xlsx")))

def load_env():
    e = {}; p = os.path.join(ROOT, ".env.local")
    if os.path.exists(p):
        for line in open(p, encoding="utf-8"):
            if "=" in line and not line.strip().startswith("#"):
                k, v = line.split("=", 1); e[k.strip()] = v.strip()
    return e
ENV = load_env()
SUPA_URL = ENV.get("VITE_SUPABASE_URL", "").rstrip("/")
SERVICE = ENV.get("SUPABASE_SERVICE_ROLE_KEY")
GKEY = ENV.get("VITE_GOOGLE_MAPS_API_KEY")

def norm(s): return re.sub(r"\s+|아파트", "", str(s or "")).strip()
def norm2(s): return norm(re.sub(r"\(.*?\)", "", str(s or "")))
def esc(s): return str(s).replace("'", "''")

cp = {}
with open(CP, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        cp.setdefault(row["gu"], {}).setdefault(norm(row["complex"]),
            {"orig": row["complex"], "n2": norm2(row["complex"]), "rows": []})["rows"].append(row)

# price_trends → 시세추이(price_history) 보강. key (gu, area_bucket) → 월별 rows(year_month 오름차순)
PT = os.path.join(ROOT, "scripts", "output", "price_trends.csv")
pt = {}
if os.path.exists(PT):
    with open(PT, encoding="utf-8") as f:
        for row in csv.DictReader(f):
            pt.setdefault((row["gu"], row["area_bucket"]), []).append(row)
    for kk in pt:
        pt[kk].sort(key=lambda r: r.get("year_month", ""))

def area_bucket(a):
    if a is None: return "미상"
    if a <= 60: return "60㎡ 이하"
    if a <= 85: return "60–85㎡"
    if a <= 102: return "85–102㎡"
    if a <= 135: return "102–135㎡"
    return "135㎡ 초과"

def price_history_for(gu, area_m2):
    out = []
    for r in pt.get((gu, area_bucket(area_m2)), []):
        ym = r.get("year_month", "")
        try: price = int(float(r.get("price")))
        except: continue
        est = str(r.get("is_estimated", "")).lower() in ("true", "t", "1")
        out.append({"month": ym[2:].replace("-", ".") if len(ym) >= 7 else ym, "yearMonth": ym, "price": price, "estimated": est})
    return out

# ---------- 단지별 개별 실거래(real-only) → 매물 상세 표 A(평형별 요약)/B(실거래 내역) ----------
CT_PATH = os.path.join(ROOT, "scripts", "output", "complex_trades.csv")
def load_complex_trades(needed):  # needed: {(gu, complex)} — 매물 단지만 메모리에
    ct = {}
    if not os.path.exists(CT_PATH): return ct
    with open(CT_PATH, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            key = (r["gu"], r["complex"])
            if key not in needed: continue
            try: ct.setdefault(key, []).append({"a": int(r["area_m2"]), "ym": r["year_month"], "d": r.get("day", ""), "fl": r.get("floor", ""), "p": int(r["price"])})
            except: pass
    return ct
def _median(xs):
    s = sorted(xs); n = len(s); return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) // 2
def area_summary(trades, my_area):  # 표 A: 단지의 평형별 요약
    by = {}
    for t in trades: by.setdefault(t["a"], []).append(t)
    out = []
    for a in sorted(by):
        ts = sorted(by[a], key=lambda x: (x["ym"], x["d"])); ps = [t["p"] for t in ts]
        out.append({"areaM2": a, "count": len(ts), "recentPrice": ts[-1]["p"], "recentMonth": ts[-1]["ym"],
                    "minPrice": min(ps), "maxPrice": max(ps), "isMine": a == my_area})
    return out
def recent_trades(trades, my_area, limit=30):  # 표 B: 내 평형 최근 실거래
    ts = sorted([t for t in trades if t["a"] == my_area], key=lambda x: (x["ym"], x["d"]), reverse=True)
    return [{"yearMonth": t["ym"], "day": t["d"], "areaM2": t["a"], "floor": t["fl"], "price": t["p"]} for t in ts[:limit]]
def real_history(trades, my_area):  # 차트: 내 평형 월별 real 중앙값 (재생산 없음)
    by = {}
    for t in trades:
        if t["a"] == my_area: by.setdefault(t["ym"], []).append(t["p"])
    return [{"month": ym[2:].replace("-", "."), "yearMonth": ym, "price": _median(by[ym]), "count": len(by[ym])} for ym in sorted(by)]

def resolve_complex(comp, gu):
    g = cp.get(gu)
    if not g: return None
    n, n2 = norm(comp), norm2(comp)
    if n in g: return g[n]
    cand = [v for v in g.values() if v["n2"] == n2 and len(n2) >= 3]
    if len(cand) == 1: return cand[0]
    sub = [v for k, v in g.items() if len(k) >= 3 and len(n2) >= 3 and (n2 in k or k in n2)]
    return sub[0] if len(sub) == 1 else None

def baseline(entry, area_m2):
    rows = entry["rows"]
    ex = [r for r in rows if int(r["area_m2"]) == area_m2 and int(r["sample_size"]) >= MIN_SAMPLE]
    if ex: return ex[0], False
    near = sorted([r for r in rows if abs(int(r["area_m2"]) - area_m2) <= 2 and int(r["sample_size"]) >= MIN_SAMPLE],
                  key=lambda r: -int(r["sample_size"]))
    return (near[0], True) if near else (None, None)

# ---------- 지오코딩 (디스크 캐시) ----------
try: _gc = json.load(open(GCACHE, encoding="utf-8"))
except Exception: _gc = {}
_calls = [0]
def _get(url):
    with urllib.request.urlopen(url, timeout=20) as r: return json.load(r)
def geocode_address(addr):
    if not (addr and GKEY): return None
    k = "addr|" + addr
    if k in _gc: return _gc[k]
    try:
        d = _get("https://maps.googleapis.com/maps/api/geocode/json?" + urllib.parse.urlencode(
            {"address": addr, "language": "ko", "region": "kr", "key": GKEY}))
        res = d.get("results", []); loc = res[0]["geometry"]["location"] if res else None
        out = {"lat": round(loc["lat"], 6), "lng": round(loc["lng"], 6)} if loc else None
    except Exception: out = None
    _gc[k] = out; _calls[0] += 1; time.sleep(0.08); return out
def places_complex(gu, comp):
    if not (comp and GKEY): return None
    k = "cx|" + gu + "|" + comp
    if k in _gc: return _gc[k]
    short = gu.split()[-1] if gu else ""
    try:
        d = _get("https://maps.googleapis.com/maps/api/place/textsearch/json?" + urllib.parse.urlencode(
            {"query": f"{short} {comp}".strip(), "language": "ko", "region": "kr", "key": GKEY}))
        res = d.get("results", []); nc = norm(comp)
        cand = [r for r in res if nc and (nc in norm(r.get("name", "")) or norm(r.get("name", "")) in nc)]
        pick = (cand or res)[0] if res else None
        loc = pick["geometry"]["location"] if pick else None
        out = {"lat": round(loc["lat"], 6), "lng": round(loc["lng"], 6)} if loc else None
    except Exception: out = None
    _gc[k] = out; _calls[0] += 1; time.sleep(0.08); return out
def jitter(idv, base):
    h = int(hashlib.md5(idv.encode()).hexdigest(), 16)
    return {"lat": round(base["lat"] + ((h % 1000) / 1000 - 0.5) * 0.03, 6),
            "lng": round(base["lng"] + (((h // 1000) % 1000) / 1000 - 0.5) * 0.03, 6)}
def resolve_coord(it, idv):
    c = places_complex(it["gu"], it["user_comp"])  # 한국 아파트는 단지명 Places 가 가장 정확
    if c: return c, "places"
    road = str(it.get("road") or it.get("legacy_addr") or "").strip()
    if road:
        c = geocode_address(road)
        if c: return c, "road"
    cen = geocode_address(it["gu"]) or {"lat": 37.8, "lng": 128.1}
    return jitter(idv, cen), "centroid"

# 생활권(주변 편의시설) — 라이브 /api/lookup-lifestyle 재사용(정상 등록과 동일 로직). 좌표 캐시.
LIFE_BASE = "https://guepmae.vercel.app/api/lookup-lifestyle"
def lifestyle_for(lat, lng):
    k = f"life|{round(lat, 5)}|{round(lng, 5)}"
    if k in _gc: return _gc[k] or {}
    out = {}
    try:
        u = LIFE_BASE + "?" + urllib.parse.urlencode({"lat": lat, "lng": lng})
        out = (json.load(urllib.request.urlopen(u, timeout=60)).get("lifestyle")) or {}
    except Exception as e:
        print("lifestyle 경고:", str(e)[:100])
    _gc[k] = out; _calls[0] += 1; time.sleep(0.15); return out

# ---------- 엑셀 읽기 (여러 파일) ----------
from openpyxl import load_workbook
def to_int(v):
    try: return int(float(str(v).replace(",", "")))
    except: return None
def floor_area(v):
    try: return int(float(v))
    except: return None
def rooms_of(a): return 1 if a and a < 40 else 2 if a and a < 60 else 3 if a and a < 95 else 4 if a else 3
def baths_of(rm, a): return 1 if rm <= 2 else (2 if (a or 0) > 80 else 1) if rm == 3 else 2
def clean_addr(a): return re.sub(r"\s*\d+동$", "", str(a or "").strip())

FIELDS = {"title": "제목", "floor": "층", "rooms": "방수", "bathrooms": "욕실수", "built_year": "준공연도",
          "direction": "향", "supply_area": "공급면적(㎡)", "maintenance_fee": "관리비(원)", "parking": "주차",
          "move_in_date": "입주가능일", "occupancy_status": "거주상태", "unit_count": "세대수",
          "description": "매물설명", "road": "도로명주소", "legacy_addr": "주소"}

kept = []; stat = collections.Counter(); per_region = collections.Counter()
for fp in FILES:
    wb = load_workbook(fp, data_only=True)
    ws = wb["매물입력"] if "매물입력" in wb.sheetnames else wb[wb.sheetnames[0]]
    H = {(ws.cell(1, i).value or "").strip(): i for i in range(1, ws.max_column + 1)}
    def gv(r, *labels, _H=H, _ws=ws):
        for lb in labels:
            for h, i in _H.items():
                if h == lb or h.rstrip("*") == lb:
                    v = _ws.cell(r, i).value
                    if v is not None and str(v).strip() != "": return v
        return None
    photo_cols = [i for h, i in H.items() if h.startswith("사진")]
    for r in range(2, ws.max_row + 1):
        comp = gv(r, "단지명")
        if not comp: continue
        stat["total"] += 1
        gu = str(gv(r, "구/시군구") or "").strip()
        area_m2 = floor_area(gv(r, "전용면적(㎡)")); price = to_int(gv(r, "호가(원)"))
        if not price or area_m2 is None: stat["bad"] += 1; continue
        entry = resolve_complex(str(comp).strip(), gu)
        if not entry: stat["no_complex"] += 1; continue
        base, approx = baseline(entry, area_m2)
        if not base: stat["no_area"] += 1; continue
        bp = int(base["median_price"]); disc = round((bp - price) / bp * 100, 1)
        if disc < MIN_DISC: stat["below"] += 1; continue
        if disc > MAX_DISC: stat["outlier"] += 1; continue
        photos = []
        for ci in photo_cols:
            v = ws.cell(r, ci).value
            if v:
                for u in re.split(r"[;\n,]", str(v)):
                    u = u.strip()
                    if u.startswith("http"): photos.append(u)
        it = {"user_comp": str(comp).strip(), "comp": entry["orig"], "gu": gu, "area_m2": area_m2,
              "area_raw": to_int(gv(r, "전용면적(㎡)")) or area_m2, "price": price, "base": base,
              "disc": disc, "approx": approx, "photos": photos}
        for key, lb in FIELDS.items():
            it[key] = gv(r, lb)
        kept.append(it); per_region[gu] += 1
    wb.close()

gus = sorted(set(it["gu"] for it in kept))
# 매물 단지의 개별 실거래(표 A/B)만 메모리 로드
CT_DATA = load_complex_trades(set((it["gu"], it["comp"]) for it in kept))

def build_row(it):
    base = it["base"]; area = it["area_raw"]
    trades = CT_DATA.get((it["gu"], it["comp"]), [])
    idv = "gm-" + hashlib.md5(f"{it['gu']}|{it['user_comp']}|{it['area_m2']}|{it['price']}|{it.get('road') or it.get('legacy_addr') or ''}".encode()).hexdigest()[:10]
    coord, csrc = resolve_coord(it, idv)
    life = lifestyle_for(coord["lat"], coord["lng"])
    sample = int(base["sample_size"]); period = f"{base.get('earliest_year_month','')}~{base.get('latest_year_month','')}".strip("~")
    title = it.get("title") or f"{it['user_comp']} 전용{it['area_m2']}㎡"
    addr = clean_addr(it.get("road") or it.get("legacy_addr") or it["gu"]) or it["gu"]
    rooms = to_int(it.get("rooms")) or rooms_of(area)
    media = [{"src": u, "label": "대표 사진" if j == 0 else f"사진 {j+1}", "alt": f"{title} 사진 {j+1}"} for j, u in enumerate(it["photos"])]
    return {
        "id": idv, "title": title, "address": addr, "coordinates": coord, "region": it["gu"],
        "property_type": "아파트", "price": it["price"], "actual_transaction_price": int(base["median_price"]),
        "discount_rate": it["disc"],
        "price_basis": {"source": "complex", "baselinePrice": int(base["median_price"]), "areaM2": int(base["area_m2"]),
                        "requestedAreaM2": it["area_m2"], "approxArea": bool(it["approx"]), "sampleSize": sample,
                        "periodStart": base.get("earliest_year_month"), "periodEnd": base.get("latest_year_month"),
                        "confidence": "high" if sample >= 5 else "medium",
                        "method": f"동일 단지 {base['area_m2']}㎡ · {period} {sample}건 중앙값", "computedAt": TODAY, "coordSource": csrc},
        "urgent_score": min(99, round(50 + it["disc"] * 3)), "area": area,
        "supply_area": to_int(it.get("supply_area")) or round(area * 1.33), "floor": it.get("floor"),
        "direction": it.get("direction"), "occupancy_status": it.get("occupancy_status"),
        "built_year": to_int(it.get("built_year")) or to_int(base.get("built_year")), "verified": True,
        "last_verified_at": TODAY, "recent_transaction_date": (base.get("latest_year_month") or "2026-06") + "-01",
        "description": it.get("description") or f"{it['gu']} 실거래가 대비 {it['disc']}% 저렴한 급매물입니다.",
        "parking": it.get("parking") or "미공개", "maintenance_fee": to_int(it.get("maintenance_fee")),
        "move_in_date": it.get("move_in_date") or "협의", "rooms": rooms,
        "bathrooms": to_int(it.get("bathrooms")) or baths_of(rooms, area), "unit_count": to_int(it.get("unit_count")),
        "agent": {"name": "급매 운영팀", "office": "급매 검증", "phone": "", "email": "", "verified": True},
        "lifestyle": life, "price_history": real_history(trades, it["area_m2"]), "media": media,
        "price_table": {
            "complexName": it["comp"], "myAreaM2": it["area_m2"], "basisPeriod": "최근 3년 실거래",
            "areaSummary": area_summary(trades, it["area_m2"]),
            "recentTrades": recent_trades(trades, it["area_m2"]),
        },
    }

rows = [build_row(it) for it in sorted(kept, key=lambda x: -x["disc"])]
# id 중복(동일 단지·면적·가격·주소) 제거 — 첫 건 유지
seen = set(); rows = [r for r in rows if not (r["id"] in seen or seen.add(r["id"]))]
json.dump(_gc, open(GCACHE, "w", encoding="utf-8"), ensure_ascii=False)

# ---------- SQL 파일(항상) ----------
def Sv(v): return "null" if v is None or v == "" else "'" + esc(v) + "'"
def Nv(v): return "null" if v is None else str(v)
def Jv(v): return "'" + esc(json.dumps(v, ensure_ascii=False)) + "'::jsonb"
JSONB = {"coordinates", "price_basis", "agent", "lifestyle", "price_history", "media", "price_table"}
NUM = {"price", "actual_transaction_price", "discount_rate", "urgent_score", "area", "supply_area",
       "built_year", "maintenance_fee", "rooms", "bathrooms", "unit_count"}
BOOL = {"verified"}
sql = io.StringIO()
sql.write("-- 급매 매물 자동 생성 (import-listings.py). SQL Editor 붙여넣기 또는 --apply.\n")
sql.write("-- 해당 지역 교체하려면 아래 delete 들 주석 해제:\n")
for gu in gus:
    sql.write(f"-- delete from public.properties where region = '{esc(gu)}';\n")
sql.write("\n")
if rows:
    keys = list(rows[0].keys())
    for row in rows:
        vals = [Jv(row[k]) if k in JSONB else ("true" if row[k] else "false") if k in BOOL
                else Nv(row[k]) if k in NUM else Sv(row[k]) for k in keys]
        sql.write(f"insert into public.properties ({', '.join(keys)}) values ({', '.join(vals)});\n")
open(OUT_SQL, "w", encoding="utf-8").write(sql.getvalue())

# ---------- Supabase 등록(--apply) ----------
applied = 0; apply_msg = "dry-run (SQL 파일만 생성)"
if APPLY:
    if not (SUPA_URL and SERVICE):
        apply_msg = "❌ --apply 인데 .env.local 에 SUPABASE_SERVICE_ROLE_KEY 없음 → 등록 건너뜀(SQL 은 생성됨)."
    else:
        hdr = {"apikey": SERVICE, "Authorization": "Bearer " + SERVICE, "Content-Type": "application/json"}
        for gu in gus:
            # 해당 지역 매물 id 수집 → 리포트(child) 먼저 삭제(FK) → 매물 삭제. 재등록 시 리포트는 새 데이터로 재생성됨.
            try:
                gid = urllib.request.Request(f"{SUPA_URL}/rest/v1/properties?region=eq.{urllib.parse.quote(gu)}&select=id", headers=hdr)
                ids = [r["id"] for r in json.load(urllib.request.urlopen(gid, timeout=30))]
            except Exception: ids = []
            if ids:
                inlist = "(" + ",".join(ids) + ")"
                rdel = urllib.request.Request(f"{SUPA_URL}/rest/v1/property_reports?property_id=in.{inlist}",
                                              method="DELETE", headers={**hdr, "Prefer": "return=minimal"})
                try: urllib.request.urlopen(rdel, timeout=30).read()
                except Exception as e: print("report DELETE 경고:", str(e)[:120])
            pdel = urllib.request.Request(f"{SUPA_URL}/rest/v1/properties?region=eq.{urllib.parse.quote(gu)}",
                                          method="DELETE", headers={**hdr, "Prefer": "return=minimal"})
            try: urllib.request.urlopen(pdel, timeout=30).read()
            except Exception as e: print("DELETE 경고:", str(e)[:120])
        ok = True
        for i in range(0, len(rows), 100):
            chunk = rows[i:i+100]
            req = urllib.request.Request(f"{SUPA_URL}/rest/v1/properties?on_conflict=id",
                data=json.dumps(chunk, ensure_ascii=False).encode("utf-8"), method="POST",
                headers={**hdr, "Prefer": "resolution=merge-duplicates,return=minimal"})
            try: urllib.request.urlopen(req, timeout=90).read(); applied += len(chunk)
            except urllib.error.HTTPError as e:
                apply_msg = f"❌ 등록 실패: {e.code} {e.read().decode('utf-8','ignore')[:200]}"; ok = False; break
        if ok: apply_msg = f"✅ Supabase 등록 완료: {applied}건 (지역 {len(gus)}개 교체)"

# ---------- 리포트 ----------
rep = io.StringIO()
rep.write("=== 매물 import 리포트 (강원 전체) ===\n")
rep.write(f"입력 파일 {len(FILES)}개, 지역 {len(gus)}개\n")
rep.write(f"엑셀 총행 {stat['total']} | 단지미매칭 {stat['no_complex']} | 표본부족 {stat['no_area']} | 일반/프리미엄(<5%) {stat['below']} | 이상치 {stat['outlier']}\n")
rep.write(f">>> 급매 채택: {len(rows)}건 (중복 제거 후)\n")
rep.write(f"지오코딩 API 신규 호출: {_calls[0]}회 (나머지는 캐시)\n")
csrc = collections.Counter(r["price_basis"]["coordSource"] for r in rows)
rep.write(f"좌표 출처: {dict(csrc)}\n")
rep.write(f"적용: {apply_msg}\n\n[지역별 급매 수]\n")
reg = collections.Counter(r["region"] for r in rows)
for gu, c in reg.most_common():
    rep.write(f"  {gu}: {c}\n")
rep.write("\n[상위 급매 20]\n")
for r in rows[:20]:
    rep.write(f"  {r['discount_rate']:>5}%  {r['title']}  호가 {r['price']:,}/기준 {r['actual_transaction_price']:,}  사진{len(r['media'])}  [{r['price_basis']['coordSource']}]\n")
open(OUT_REP, "w", encoding="utf-8").write(rep.getvalue())
print(f"급매 {len(rows)}건 | 지오코딩 신규 {_calls[0]}회 | {apply_msg}")
