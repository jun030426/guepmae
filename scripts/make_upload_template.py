# 매물 업로드 엑셀 템플릿 생성기 (v3 — 사진 다중 칸 + 도로명주소 + 교훈 반영)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

FONT = "맑은 고딕"
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "매물_업로드_템플릿.xlsx")

# key, label, group, format, example, note, width
cols = [
    ("complex_name", "단지명", "필수", "텍스트", "헬리오시티",
     "★호갱노노 정식 단지명 그대로(괄호·차수 포함). 국토부 표기와 다르면 같은 단지 기준가 매칭이 빗나가 일부 매물이 '구 평균'으로 빠지거나 제외됨.", 18),
    ("gu", "구/시군구", "필수", "텍스트", "서울특별시 송파구",
     "'시도 + 시군구' 정식명칭. 예: 서울특별시 강남구 / 경기도 성남시 / 부산광역시 해운대구 / 강원특별자치도 춘천시.", 18),
    ("area_m2", "전용면적(㎡)", "필수", "숫자", "84.97",
     "전용면적(공급면적 아님). 소수점 그대로.", 13),
    ("price", "호가(원)", "필수", "숫자(원)", "1742000000",
     "실제 매물 호가. 원 단위 숫자만. 호가는 보통 실거래가보다 비쌈 → 진짜 급매(5%↓)는 소수지만 넓게 모으면 시스템이 골라냄.", 16),
    ("road_address", "도로명주소", "권장", "텍스트", "서울특별시 송파구 송파대로 345",
     "★정식 도로명/지번 주소. 지도 핀·표시에 사용. 건물 동번호('429동')만 쓰면 지도가 시내에 뭉침 — 도로명주소를 넣어야 정확.", 28),
    ("floor", "층", "권장", "텍스트/숫자", "12", "'12' 또는 '12층'.", 8),
    ("rooms", "방수", "권장", "숫자", "3", "", 7),
    ("bathrooms", "욕실수", "권장", "숫자", "2", "", 8),
    ("built_year", "준공연도", "권장", "숫자(연도)", "2018", "비우면 단지 데이터에서 자동.", 10),
    ("title", "제목", "권장", "텍스트", "헬리오시티 84㎡ 급매", "비우면 단지명+면적으로 자동 생성.", 22),
    ("ref_price_observed", "참고:표시실거래가(원)", "선택", "숫자(원)", "1850000000",
     "호갱노노에 보이는 최근 실거래가(있으면). 교차검증용 — 안 넣어도 됨.", 18),
    ("direction", "향", "선택", "텍스트", "남향", "드롭다운 선택.", 10),
    ("supply_area", "공급면적(㎡)", "선택", "숫자", "113", "비우면 전용면적×1.33 추정.", 13),
    ("maintenance_fee", "관리비(원)", "선택", "숫자(원)", "250000", "", 13),
    ("parking", "주차", "선택", "텍스트", "세대당 1.3대", "", 14),
    ("move_in_date", "입주가능일", "선택", "텍스트", "즉시 입주", "", 16),
    ("occupancy_status", "거주상태", "선택", "텍스트", "공실", "드롭다운: 공실/거주중/세입자/협의.", 11),
    ("unit_count", "세대수", "선택", "숫자", "9510", "", 10),
    ("description", "매물설명", "선택", "텍스트", "역세권 대단지 급매물", "비우면 자동 문구.", 30),
    ("source_url", "출처URL", "선택", "URL", "https://hogangnono.com/apt/...", "매물 원본 링크(검증·재수집용).", 26),
    # ----- 사진: 한 칸에 한 장씩 (구분 쉽게) -----
    ("photo1", "사진1(대표)", "사진", "URL", "https://example.com/1.jpg",
     "대표 사진(첫 번째로 노출). 호갱노노 매물의 첫 사진 URL.", 30),
    ("photo2", "사진2", "사진", "URL", "https://example.com/2.jpg",
     "추가 사진. 호갱노노에서 사진을 누르면 나오는 것들을 한 칸에 하나씩.", 24),
    ("photo3", "사진3", "사진", "URL", "https://example.com/3.jpg", "", 24),
    ("photo4", "사진4", "사진", "URL", "", "", 24),
    ("photo5", "사진5", "사진", "URL", "", "", 24),
    ("photo6", "사진6", "사진", "URL", "", "6장 넘으면 알려주세요(칸 추가). 빈 칸은 무시됨.", 24),
]

GROUP_FILL = {"필수": "FAD4D0", "권장": "FCEFC7", "선택": "EDEFF2", "사진": "D6E4F5"}
thin = Side(style="thin", color="C9CDD6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "매물입력"

for i, (key, label, group, fmt, example, note, width) in enumerate(cols, start=1):
    head = ws.cell(row=1, column=i, value=label + ("*" if group == "필수" else ""))
    head.font = Font(name=FONT, bold=True, size=10, color="1A2233")
    head.fill = PatternFill("solid", fgColor=GROUP_FILL[group])
    head.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    head.border = border
    head.comment = Comment(f"[{group}] {fmt}\n예: {example}\n{note}".strip(), "급매")

    ex = ws.cell(row=2, column=i, value=example)
    ex.font = Font(name=FONT, italic=True, size=10, color="8A93A6")
    ex.alignment = Alignment(horizontal="left", vertical="center")
    ex.border = border

    ws.column_dimensions[get_column_letter(i)].width = width

ws.freeze_panes = "E2"  # 필수 4개 고정
ws.row_dimensions[1].height = 32
ws["A2"].comment = Comment("이 2행은 예시입니다. 지우고 3행부터 실제 매물을 입력하세요. (한 행 = 한 매물)", "급매")

idx = {key: i for i, (key, *_rest) in enumerate(cols, start=1)}
dv_dir = DataValidation(type="list", formula1='"남향,남동향,남서향,동향,서향,북동향,북서향,북향"', allow_blank=True)
dv_occ = DataValidation(type="list", formula1='"공실,거주중,세입자,협의"', allow_blank=True)
ws.add_data_validation(dv_dir)
ws.add_data_validation(dv_occ)
dv_dir.add(f"{get_column_letter(idx['direction'])}2:{get_column_letter(idx['direction'])}500")
dv_occ.add(f"{get_column_letter(idx['occupancy_status'])}2:{get_column_letter(idx['occupancy_status'])}500")

# ----- 작성가이드 시트 -----
g = wb.create_sheet("작성가이드")
notes = [
    "급매 매물 업로드 템플릿 v3 — 작성 가이드",
    "",
    "[기본 규칙]",
    "• '매물입력' 시트 3행부터 입력 (2행 예시는 지우고). 한 행 = 한 매물.",
    "• 빨강(필수) 4개는 반드시. 가격은 '원' 단위 숫자만(예: 1742000000). 면적은 전용면적(㎡).",
    "• 파랑(사진) 칸은 한 칸에 사진 1장씩. 사진1=대표. 빈 칸은 무시.",
    "",
    "[춘천 수집에서 배운 점 — 꼭 지키기]",
    "① 단지명: 호갱노노 정식 명칭 그대로(괄호·차수 포함). 표기가 다르면 일부 매물이 매칭 안 됨.",
    "② 도로명주소: 정식 도로명/지번 주소. '429동' 같은 건물 동번호만 있으면 지도가 시내에 뭉침.",
    "③ 구/시군구: '시도 + 시군구' 정식. 예: 서울특별시 강남구 / 경기도 성남시 / 부산광역시 해운대구.",
    "④ 호가는 보통 실거래가보다 비쌈 → 진짜 급매는 소수(춘천 715건 중 42건). 넓게 모으면 시스템이 골라냄.",
    "⑤ 사진은 실사진 우선. 여러 장은 사진1~6 칸에 하나씩.",
    "",
    "[작업 흐름]",
    "엑셀 채우기 → 담당(클로드)에게 파일 전달 → 자동으로 급매 선별+지오코딩+Supabase 등록.",
    "(엑셀을 Supabase에 직접 넣지 마세요. 스크립트가 처리합니다.)",
    "",
    "[시스템이 자동으로 채우는 것 — 넣지 마세요]",
    "기준 실거래가 · 할인율 · 산출근거 · 좌표(도로명주소로 정밀) · 지역 · 생활권 · 검증여부 · 시세추이.",
    "",
]
r = 1
for line in notes:
    c = g.cell(row=r, column=1, value=line)
    bold = line.startswith("[") or r == 1
    c.font = Font(name=FONT, bold=bold, size=(13 if r == 1 else 10), color="1A2233")
    r += 1

hdr = ["컬럼", "키(영문)", "구분", "형식", "예시", "설명"]
for j, h in enumerate(hdr, start=1):
    c = g.cell(row=r, column=j, value=h)
    c.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1A2233")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
r += 1
for (key, label, group, fmt, example, note, width) in cols:
    for j, v in enumerate([label, key, group, fmt, example, note], start=1):
        c = g.cell(row=r, column=j, value=v)
        c.font = Font(name=FONT, size=10, color="1A2233")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = border
        if j == 3:
            c.fill = PatternFill("solid", fgColor=GROUP_FILL[group])
    r += 1
for j, w in enumerate([18, 20, 8, 14, 28, 60], start=1):
    g.column_dimensions[get_column_letter(j)].width = w

wb.save(OUT)
print("SAVED:", OUT, "| cols", len(cols))
